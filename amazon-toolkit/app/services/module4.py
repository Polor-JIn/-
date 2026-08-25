"""模块 4：1688 映射维护 + 订单匹配 + 采购单生成"""
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from app import config, db
from app.services import sku_gen


def update_mapping(product_id, p1688_id="", url="", spec=""):
    conn = db.get_conn()
    conn.execute(
        "UPDATE product_master SET p1688_id=?, purchase_url=?, purchase_spec=?, updated_at=? WHERE product_id=?",
        (str(p1688_id).strip(), str(url).strip(), str(spec).strip(), db.now(), product_id))
    conn.commit()


def match_orders(df, mapping, batch: str):
    """上传订单 → 按 SKU 匹配 1688 映射，生成匹配结果（未映射的列出供补录）"""
    conn = db.get_conn()
    matched, unmatched = [], []
    for _i, row in df.iterrows():
        def g(f):
            col = mapping.get(f)
            v = row[col] if col else None
            return str(v).strip() if v is not None else ""

        sku = g("sku")
        qty = _int(g("qty")) or 1
        product = sku_gen.resolve(sku)
        if product and product["p1688_id"]:
            matched.append({
                "p1688_id": product["p1688_id"], "url": product["purchase_url"] or "",
                "spec": product["purchase_spec"] or "", "sku": sku, "qty": qty,
                "unit_price": product["purchase_price"] or "",
                "address": db.get_config("purchase_address", ""), "product_id": product["product_id"],
            })
        else:
            unmatched.append({"sku": sku, "qty": qty,
                              "reason": "未匹配1688映射" if product else "产品未入库"})
    return {"matched": matched, "unmatched": unmatched, "batch": batch}


def save_purchase_orders(matched, batch: str):
    conn = db.get_conn()
    for m in matched:
        conn.execute(
            "INSERT INTO purchase_orders(batch_id, p1688_id, url, sku, spec, qty, unit_price, address, created_at) "
            "VALUES(?,?,?,?,?,?,?,?,?)",
            (batch, m["p1688_id"], m["url"], m["sku"], m["spec"], m["qty"],
             _f(m["unit_price"]) or None, m["address"], db.now()))
    conn.commit()
    conn.execute("INSERT OR IGNORE INTO batch_log(batch_type, filename, batch_id, summary, created_at) "
                 "VALUES('purchase', ?, ?, ?, ?)", (batch, batch, f"orders={len(matched)}", db.now()))
    conn.commit()
    return len(matched)


def write_purchase_xlsx(matched, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "1688采购单"
    headers = ["ID(1688)", "1688链接", "亚马逊子SKU", "1688规格", "数量", "采购单价", "采购地址"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color="FCE4D6")
    for i, m in enumerate(matched, 2):
        ws.cell(i, 1, m["p1688_id"])
        ws.cell(i, 2, m["url"])
        ws.cell(i, 3, m["sku"])
        ws.cell(i, 4, m["spec"])
        ws.cell(i, 5, m["qty"])
        ws.cell(i, 6, m["unit_price"])
        ws.cell(i, 7, m["address"])
    for c, w in zip("ABCDEFG", [14, 42, 24, 22, 8, 10, 30]):
        ws.column_dimensions[c].width = w
    ws.freeze_panes = "A2"
    wb.save(path)
    return path


def current_batch() -> str:
    return f"P{uuid.uuid4().hex[:8]}"


def _int(v):
    try:
        return int(float(str(v).strip()))
    except (TypeError, ValueError):
        return 1


def _f(v):
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return None