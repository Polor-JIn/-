"""模块 1：跟卖数据清洗 → SKU 生成 → 去重入库 → 上架 Excel 生成"""
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app import config, db
from app.services import dicts, sku_gen, dedup


def process_listing(df, mapping, batch: str):
    """处理跟卖数据。返回 (summary, out_path, pending_list)"""
    conn = db.get_conn()
    summary = {"added": 0, "skipped": 0, "conflict": 0, "pending": 0,
               "new_colors": set(), "new_sizes": set()}
    rows = []

    for _i, row in df.iterrows():
        def g(f):
            col = mapping.get(f)
            v = row[col] if col else None
            return str(v).strip() if v is not None else ""

        asin, title = g("asin"), g("title")
        color_raw, size_raw = g("color"), g("size")

        color_std, cs = dicts.normalize(color_raw, "color")
        size_std, ss = dicts.normalize(size_raw, "size")
        if cs == "new":
            color_std = "*" + color_std
            summary["new_colors"].add(color_raw)
        if ss == "new":
            size_std = "*" + size_std
            summary["new_sizes"].add(size_raw)
        if not color_std or not size_std or color_std.startswith("*") or size_std.startswith("*"):
            rows.append({"sku": "", "asin": asin, "title": title, "color": color_raw,
                         "size": size_raw, "action": "待确认颜色/尺码", "price": ""})
            summary["pending"] += 1
            continue

        sku = sku_gen.build_sku(color_std, size_std, title)
        res = dedup.ingest(batch, sku, {"asin": asin, "title": title,
                                        "color": color_std, "size": size_std}, source="跟卖")
        action = {"added": "新增SKU", "skipped": "已存在-跳过", "conflict": "冲突-待人工"}[res["action"]]
        summary[res["action"]] += 1
        rows.append({"sku": sku, "asin": asin, "title": title, "color": color_std,
                     "size": size_std, "action": action, "price": g("price")})

    out_path = str(config.OUTPUT_DIR / f"上架表_{batch}.xlsx")
    _write_listing_xlsx(out_path, rows)
    summary["total"] = len(rows)
    return summary, out_path, (summary["new_colors"], summary["new_sizes"])


def _write_listing_xlsx(path, rows):
    """按用户 '跟卖商品模板' 的 8 列结构导出：
    Asin | SKU | 价格 | 采购链接 | 保底价 | 单次调整金额 | 库存 | 最大订单数量
    价格 / 保底价 留空高亮（待填）；采购链接取商品库已录的 1688 链接。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "跟卖商品"
    headers = ["Asin", "SKU", "价格", "采购链接", "保底价",
               "单次调整金额", "库存", "最大订单数量"]
    header_fill = PatternFill("solid", start_color="4472C4")
    pending_fill = PatternFill("solid", start_color="FFF2CC")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    inv = db.get_config("inventory", "10")
    adj = db.get_config("adjust_price", "0.50")
    maxp = db.get_config("max_purchase", "5")

    for i, r in enumerate(rows, start=2):
        ws.cell(i, 1, r["asin"])
        ws.cell(i, 2, r["sku"])
        price_cell = ws.cell(i, 3)   # 价格：留空待填，黄底
        floor_cell = ws.cell(i, 5)   # 保低价：留空待填，黄底
        for cell in (price_cell, floor_cell):
            cell.fill = pending_fill
        # 采购链接：优先取商品库已维护的 1688 链接
        conn = db.get_conn()
        row = conn.execute("SELECT purchase_url FROM product_master WHERE sku=?", (r["sku"],)).fetchone()
        ws.cell(i, 4, row["purchase_url"] if row and row["purchase_url"] else "")
        ws.cell(i, 6, adj)
        ws.cell(i, 7, inv)
        ws.cell(i, 8, maxp)
    widths = [16, 20, 10, 42, 10, 16, 8, 14]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + c)].width = w
    ws.freeze_panes = "A2"
    wb.save(path)


def make_batch(prefix: str = "") -> str:
    return f"{prefix}{uuid.uuid4().hex[:8]}"