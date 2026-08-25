"""利润精算 + 保底价公式 + 利润分析报告"""
from app import db
from app.services import freight as freight_svc, weight as weight_svc


def floor_price(product, conn=None):
    """保底价 = 成本公式反推：(采购成本 + 预估运费) / (1 - 佣金率 - 目标毛利)
    重量使用优先级逻辑的有效重量；无重量则不返回数值。返回 (floor, weight, source, note)"""
    conn = conn or db.get_conn()
    cost = product["purchase_price"]
    if not cost:
        return None, None, "无", "缺采购成本"
    w, wsrc = weight_svc.effective_weight(product["product_id"])
    if not w:
        return None, w, wsrc, "缺重量，无法估算"
    comm = float(db.get_config("commission_rate", "0.15")) if db.get_config("commission_rate", "") else 0.15
    margin = product["target_margin"] if product["target_margin"] is not None else \
        float(db.get_config("target_margin", "0.30"))
    freight_est, add, _ = freight_svc.total_freight("", w, "")
    freight_est = freight_est or 0.0
    denom = 1 - comm - margin
    if denom <= 0:
        return None, w, wsrc, "佣金+目标毛利>=100%，参数错误"
    floor = (cost + freight_est + add) / denom
    return round(floor, 2), w, wsrc, "预估" if wsrc in ("历史均值", "参考重量", "无") else "确认重量"


def net_profit(price, cost, freight_total, commission_rate):
    """单件净利 = 售价*(1-佣金) - 采购成本 - 运费(含附加费)"""
    if not price:
        return None
    return round(price * (1 - commission_rate) - (cost or 0) - (freight_total or 0), 2)


def report(conn=None):
    """按产品维度出利润现状（精算口径），供调价参考。返回 (rows, meta)"""
    conn = conn or db.get_conn()
    comm = float(db.get_config("commission_rate", "0.15"))
    products = conn.execute("SELECT * FROM product_master").fetchall()
    rows = []
    for p in products:
        w, wsrc = weight_svc.effective_weight(p["product_id"])
        floor, fw, _s, note = floor_price(
            {"product_id": p["product_id"], "purchase_price": p["purchase_price"], "target_margin": p["target_margin"]}, conn)
        freight_est = 0.0
        if w:
            f, add, _ = freight_svc.total_freight("", w, "", conn)
            freight_est = f or 0.0
        latest = conn.execute(
            "SELECT * FROM price_history WHERE product_id=? ORDER BY ref_date DESC, id DESC LIMIT 1",
            (p["product_id"],)).fetchone()
        rows.append({
            "sku": p["sku"], "asin": p["asin"] or "", "title": (p["title"] or "")[:40],
            "color": p["color"] or "", "size": p["size"] or "",
            "purchase_price": p["purchase_price"] or 0,
            "weight": w or 0, "weight_source": wsrc,
            "freight_est": freight_est,
            "price": latest["price"] if latest else None,
            "price_source": latest["price_source"] if latest else "",
            "floor_price": floor, "floor_note": note,
            "net_profit": net_profit(latest["price"] if latest else None,
                                     p["purchase_price"], freight_est, comm),
            "commission_rate": comm,
        })
    rows = [r for r in rows if r["price"] is not None or r["floor_price"] is not None]
    return rows


def write_report_xlsx(rows, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "利润分析"
    headers = ["SKU", "ASIN", "标题", "颜色", "尺码", "采购成本", "重量(kg)", "重量来源",
               "预估运费", "当前售价", "售价来源", "保底价", "保底说明", "单件净利", "佣金率"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color="D9E1F2")
    for i, r in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            ws.cell(i, c, r.get(h))
    dims = [18, 14, 46, 10, 8, 10, 10, 10, 10, 10, 10, 10, 14, 10, 8]
    for c, w in enumerate(dims, 1):
        ws.column_dimensions[chr(64 + c)].width = w
    ws.freeze_panes = "A2"
    wb.save(path)
    return path